import time
from selenium import webdriver
import openpyxl

wb = openpyxl.Workbook()

PATH = (r"C:\webdrivers\geckodriver.exe")
driver = webdriver.Firefox(executable_path=PATH)

driver.get("https://web.whatsapp.com/")


def numbers_collecter_by_group_link(x):
    #Search by group link
    try:
        url = x
        driver.get(url)
        time.sleep(10)
    except:
        print("Invite link of group is wrong.")
        pass
    try:
        #Join Group
        driver.find_element_by_xpath("""/html/body/div[1]/div/span[2]/div/div/div/div/div/div/div[2]/div/div[2]/div/div""").click()
        time.sleep(3)
    except:
        print("You have already participated in this group!!!====>",url)
        pass
    #Get group name
    driver.find_element_by_xpath("""/html/body/div[1]/div/div/div[4]/div/header/div[1]""").click() #profile click
    time.sleep(2)
    name = driver.find_element_by_xpath("""/html/body/div[1]/div/div/div[2]/div[3]/span/div/span/div/div/section/div[1]/div/div[2]/div/div/div[1]/div/div[2]""").text

    #Get Group Creation info
    try:
        desc = driver.find_element_by_xpath("""/html/body/div[1]/div/div/div[2]/div[3]/span/div/span/div/div/section/div[2]/div""").text
        print(desc)
    except:
        desc = 'null'
    #Get Group members
    numbers = driver.find_element_by_xpath("""/html/body/div[1]/div/div/div[4]/div/header/div[2]/div[2]/span""").text
    list_numbers = numbers.split(',')
    print(list_numbers)

    #Create sheet with group name
    ws = wb.create_sheet()
    ws.title = name
    group_info_excel = ws['A1'] 
    group_info_excel.value = desc
    group_url_excel = ws['A2']
    group_url_excel.value = url

    for r in range(0,len(list_numbers)):
        ws.cell(row=r+1,column=2).value=list_numbers[r]


links = input('Enter Group Links with space: ')
links = links.split()
groups_list = ['https://web.whatsapp.com/accept?code='+ x[-22:] for x in links]

for x in groups_list:
    numbers_collecter_by_group_link(x)
    wb.save("data.xlsx")
